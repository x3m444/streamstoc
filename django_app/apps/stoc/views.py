from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import connection, IntegrityError, InternalError
from django.http import HttpResponse
from datetime import date, datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEFAULT_SHIP = 978
DEFAULT_TRAGATOR = "COSTEL"


def _full_access_required(view_func):
    """Decorator: utilizatorii read-only sunt redirectați la rapoarte."""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.conf import settings
            return redirect(settings.LOGIN_URL)
        if getattr(request.user, 'is_readonly', False):
            messages.warning(request, "Contul tău are acces doar la Rapoarte.")
            return redirect('stoc:rapoarte')
        return view_func(request, *args, **kwargs)
    return wrapper

ALL_FIELDS = [
    "Nava", "Nr Lista", "ID Lista", "Locatie", "Lungime",
    "Nr Cabluri", "Dosar", "Tragator", "Data", "Data trimisa",
    "Trimis", "Rework", "Data Rework", "Ore Rework"
]

FIELD_TYPES = {
    "Nava": "numeric", "Nr Lista": "text", "ID Lista": "text",
    "Locatie": "text", "Lungime": "numeric", "Nr Cabluri": "numeric",
    "Dosar": "text", "Tragator": "text", "Data": "date",
    "Data trimisa": "date", "Trimis": "boolean", "Rework": "boolean",
    "Data Rework": "date", "Ore Rework": "numeric"
}


def _db_error_msg(e):
    msg = str(e)
    if 'Nr Lista' in msg and 'unique' in msg.lower():
        return "Nr Lista există deja pentru această navă! Alege alt număr."
    if 'ID Lista' in msg and 'unique' in msg.lower():
        return "ID Lista există deja pentru această navă! Alege alt ID."
    # Trigger exceptions contain Romanian messages after the last newline
    if 'STOP' in msg or 'EROARE' in msg or 'REWORK' in msg.upper():
        last_line = [l.strip() for l in msg.splitlines() if l.strip()][-1]
        return last_line
    return f"Eroare bază de date: {msg}"


def _get_locatii(nava):
    with connection.cursor() as cursor:
        cursor.execute(
            'SELECT DISTINCT "Locatie" FROM list963 WHERE "Nava" = %s AND "Locatie" IS NOT NULL ORDER BY "Locatie"',
            [nava]
        )
        return [row[0] for row in cursor.fetchall()]


def _fmt_val(v):
    if isinstance(v, (date, datetime)):
        return v.strftime('%d.%m.%y')
    return v


def _rows_to_dicts(cursor, cols):
    return [dict(zip(cols, row)) for row in cursor.fetchall()]


def _make_excel(rows, headers, title, summary=None, orientation='auto'):
    """Format identic cu add_export_buttons din Streamlit."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    DATA_SIZE = 13
    HDR_HEIGHT = 24
    ROW_HEIGHT = 22

    def wcell(r, c, val='', fill_hex=None, bold=False, size=DATA_SIZE, num_fmt=None):
        cl = ws.cell(row=r, column=c, value=val)
        cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        cl.border = border
        cl.font = Font(bold=bold, size=size)
        if fill_hex:
            cl.fill = PatternFill('solid', fgColor=fill_hex)
        if num_fmt:
            cl.number_format = num_fmt
        return cl

    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    start_row = 6 if summary else 3

    # Rândurile 1-2: titlu
    ws.merge_cells(f'A1:{last_col}2')
    t = ws['A1']
    t.value = title.replace('_', ' ').upper()
    t.font = Font(bold=True, size=18)
    t.alignment = Alignment(horizontal='center', vertical='center')
    t.fill = PatternFill('solid', fgColor='D9D9D9')
    t.border = border
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    # Rândurile 4-5: sumar (dacă există)
    if summary:
        labels = ['METRI', 'CABLURI', 'LISTE']
        values = [summary.get('metri', 0), summary.get('cabluri', 0), summary.get('liste', 0)]
        for i, (lbl, val) in enumerate(zip(labels, values), 1):
            wcell(4, i, lbl, fill_hex='F2F2F2', bold=True, size=DATA_SIZE)
            wcell(5, i, val, bold=True, size=DATA_SIZE + 2, num_fmt='#,##0')
        ws.row_dimensions[4].height = HDR_HEIGHT
        ws.row_dimensions[5].height = HDR_HEIGHT

    # Header tabel
    hdr_row = start_row + 1
    for i, h in enumerate(headers, 1):
        wcell(hdr_row, i, h, fill_hex='BFBFBF', bold=True, size=DATA_SIZE)
    ws.row_dimensions[hdr_row].height = HDR_HEIGHT

    # Date
    for r_idx, row in enumerate(rows, hdr_row + 1):
        for c_idx, val in enumerate(row, 1):
            wcell(r_idx, c_idx, val)
        ws.row_dimensions[r_idx].height = ROW_HEIGHT

    # Lățimi coloane auto
    min_w = 25 if num_cols <= 5 else 12
    summary_labels = ['TOTAL METRI', 'TOTAL CABLURI', 'TOTAL LISTE'] if summary else []
    for i, h in enumerate(headers, 1):
        col_vals = [str(row[i-1]) for row in rows if row[i-1] is not None]
        if i <= len(summary_labels):
            col_vals.append(summary_labels[i - 1])
        max_len = max([len(h)] + [len(v) for v in col_vals] + [0]) + 4
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len, min_w), 50)

    # Autofilter + freeze
    ws.auto_filter.ref = f'A{hdr_row}:{last_col}{hdr_row}'
    ws.freeze_panes = f'A{hdr_row + 1}'

    # Pagină A4 — auto: portrait ≤5 coloane, landscape >5
    if orientation == 'auto':
        orientation = 'portrait' if num_cols <= 5 else 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ─────────────────────────────────────────────────────────────────────────────
# Introducere Date
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def index(request):
    if getattr(request.user, 'is_readonly', False):
        return redirect('stoc:rapoarte')
    nava = int(request.session.get('last_nava', DEFAULT_SHIP))
    tragator = request.session.get('last_tragator', DEFAULT_TRAGATOR)
    today = date.today()
    # Auto-reset la today dacă data din sesiune e mai veche
    stored = request.session.get('last_data', str(today))
    try:
        stored_date = date.fromisoformat(stored)
    except ValueError:
        stored_date = today
    if stored_date < today:
        stored_date = today
        request.session['last_data'] = str(today)
    last_data = str(stored_date)

    if request.method == 'POST':
        action = request.POST.get('action', 'save')

        if action == 'settings':
            request.session['last_nava'] = int(request.POST.get('nava', DEFAULT_SHIP))
            request.session['last_tragator'] = request.POST.get('tragator', DEFAULT_TRAGATOR)
            request.session['last_data'] = request.POST.get('data_default', str(today))
            return redirect('stoc:index')

        nr_lista = request.POST.get('nr_lista', '').strip()
        id_lista = request.POST.get('id_lista', '').strip()
        locatie = request.POST.get('locatie', '').strip()
        lungime_raw = request.POST.get('lungime', '').strip()
        nr_cabluri_raw = request.POST.get('nr_cabluri', '').strip()
        data_str = last_data

        if not all([nr_lista, id_lista, locatie, lungime_raw, nr_cabluri_raw]):
            messages.error(request, "Toate câmpurile sunt obligatorii!")
        else:
            try:
                lungime = float(lungime_raw.replace(',', '.'))
                nr_cabluri = int(nr_cabluri_raw)
            except ValueError:
                messages.error(request, "Lungime și Nr Cabluri trebuie să fie numere valide!")
            else:
                try:
                    with connection.cursor() as cursor:
                        cursor.execute('''
                            INSERT INTO list963
                            ("Nava","Nr Lista","ID Lista","Locatie","Lungime","Nr Cabluri","Data","Tragator","Trimis")
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ''', [nava, nr_lista, id_lista, locatie, lungime, nr_cabluri, data_str, tragator, False])
                    messages.success(request, f"Salvat: {id_lista} — {data_str}")
                    return redirect('stoc:index')
                except Exception as e:
                    messages.error(request, _db_error_msg(e))

    locatii = _get_locatii(nava)

    with connection.cursor() as cursor:
        cursor.execute('''
            SELECT "Nr Lista","Locatie","Nr Cabluri","Lungime","ID Lista"
            FROM list963 WHERE "Data" = %s
            ORDER BY "Locatie" ASC, "Nr Lista" ASC
        ''', [last_data])
        cols = ['nr_lista', 'locatie', 'nr_cabluri', 'lungime', 'id_lista']
        today_entries = _rows_to_dicts(cursor, cols)

    total_m = sum(float(e['lungime'] or 0) for e in today_entries)
    total_c = sum(int(e['nr_cabluri'] or 0) for e in today_entries)

    return render(request, 'stoc/index.html', {
        'locatii': locatii,
        'today': today,
        'last_data': last_data,
        'today_entries': today_entries,
        'total_m': int(total_m),
        'total_c': total_c,
        'total_l': len(today_entries),
        'last_nava': nava,
        'last_tragator': tragator,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Expediție
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def set_nava(request):
    if request.method == 'POST':
        try:
            request.session['last_nava'] = int(request.POST.get('nava', 978))
        except ValueError:
            pass
    return redirect(request.POST.get('next', '/'))


@_full_access_required
def expeditie(request):
    nava = int(request.session.get('last_nava', DEFAULT_SHIP))
    today = date.today()

    if request.method == 'POST':
        ids = request.POST.getlist('selected_ids')
        dosar = request.POST.get('dosar', '').strip()
        data_exp = request.POST.get('data_expediere', str(today))

        if not dosar:
            messages.error(request, "Introdu numele dosarului!")
        elif not ids:
            messages.error(request, "Selectează cel puțin o listă!")
        else:
            try:
                ids_int = [int(i) for i in ids]
                with connection.cursor() as cursor:
                    cursor.execute('''
                        UPDATE list963 SET "Trimis"=TRUE, "Data trimisa"=%s, "Dosar"=%s
                        WHERE "ID" = ANY(%s::int[])
                    ''', [data_exp, dosar, ids_int])
                messages.success(request, f"Liste marcate pentru expediție în dosarul {dosar}!")
                return redirect(f"{request.path}?data={data_exp}")
            except Exception as e:
                messages.error(request, _db_error_msg(e))

    # Data selectată — din GET param (după redirect) sau azi
    data_sel_str = request.GET.get('data', str(today))
    try:
        data_sel = date.fromisoformat(data_sel_str)
    except ValueError:
        data_sel = today
        data_sel_str = str(today)

    with connection.cursor() as cursor:
        cursor.execute('''
            SELECT "ID","Nr Lista","Locatie","Lungime","Nr Cabluri"
            FROM list963
            WHERE ("Trimis" IS NOT TRUE) AND ("Rework" IS NOT TRUE) AND "Nava" = %s
            ORDER BY "Nr Lista" ASC
        ''', [nava])
        records = _rows_to_dicts(cursor, ['id', 'nr_lista', 'locatie', 'lungime', 'nr_cabluri'])

        cursor.execute('''
            SELECT "Nr Lista","Locatie","Lungime","Nr Cabluri","Dosar"
            FROM list963
            WHERE "Data trimisa" = %s AND "Nava" = %s
            ORDER BY "Dosar" ASC, "Locatie" ASC, "Nr Lista" ASC
        ''', [data_sel_str, nava])
        expediate_azi = _rows_to_dicts(cursor, ['nr_lista', 'locatie', 'lungime', 'nr_cabluri', 'dosar'])

    exp_metri = int(sum(float(r['lungime'] or 0) for r in expediate_azi))
    exp_cabluri = int(sum(int(r['nr_cabluri'] or 0) for r in expediate_azi))
    exp_dosare = ', '.join(sorted(set(r['dosar'] for r in expediate_azi if r['dosar'])))

    return render(request, 'stoc/expeditie.html', {
        'records': records,
        'today': today,
        'data_sel': data_sel,
        'expediate_azi': expediate_azi,
        'exp_sumar': {'liste': len(expediate_azi), 'metri': exp_metri, 'cabluri': exp_cabluri, 'dosare': exp_dosare},
    })


# ─────────────────────────────────────────────────────────────────────────────
# Rapoarte
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def rapoarte(request):
    nava = int(request.session.get('last_nava', DEFAULT_SHIP))
    with connection.cursor() as cursor:
        today = date.today()
        cursor.execute('''
            SELECT "Nr Lista","Locatie","Lungime","Nr Cabluri"
            FROM list963 WHERE "Data" = %s AND "Nava" = %s
            ORDER BY "Locatie" ASC, "Nr Lista" ASC
        ''', [today, nava])
        azi = _rows_to_dicts(cursor, ['nr_lista', 'locatie', 'lungime', 'nr_cabluri'])

        cursor.execute('''
            SELECT COALESCE(SUM("Lungime"),0), COALESCE(SUM("Nr Cabluri"),0), COUNT(*)
            FROM list963 WHERE "Data" >= DATE_TRUNC('week', %s::date) AND "Nava" = %s
        ''', [today, nava])
        row = cursor.fetchone()
        sapt_sumar = {'metri': int(row[0]), 'cabluri': int(row[1]), 'liste': int(row[2])}

        cursor.execute('''
            SELECT "Data", SUM("Lungime"), SUM("Nr Cabluri"), COUNT(*)
            FROM list963 WHERE "Nava" = %s GROUP BY "Data" ORDER BY "Data" DESC
        ''', [nava])
        zile = _rows_to_dicts(cursor, ['data', 'metri', 'cabluri', 'liste'])

        cursor.execute('''
            SELECT COALESCE(SUM("Lungime"),0), COALESCE(SUM("Nr Cabluri"),0), COUNT(*)
            FROM list963 WHERE DATE_TRUNC('month',"Data") = DATE_TRUNC('month',%s::date) AND "Nava" = %s
        ''', [today, nava])
        row = cursor.fetchone()
        luna_sumar = {'metri': int(row[0]), 'cabluri': int(row[1]), 'liste': int(row[2])}

        cursor.execute('''
            SELECT EXTRACT(YEAR FROM "Data")::INT, EXTRACT(WEEK FROM "Data")::INT,
                   SUM("Lungime"), SUM("Nr Cabluri"), COUNT(*)
            FROM list963 WHERE "Nava" = %s GROUP BY 1,2 ORDER BY 1 DESC, 2 DESC
        ''', [nava])
        saptamani = _rows_to_dicts(cursor, ['an', 'nr_sapt', 'metri', 'cabluri', 'liste'])

        cursor.execute('''
            SELECT "Nr Lista","ID Lista","Locatie","Lungime","Nr Cabluri",
                   "Data","Tragator","Trimis","Data trimisa","Nava","Dosar","Rework","Data Rework","Ore Rework"
            FROM list963 WHERE "Nava" = %s ORDER BY "Nr Lista" ASC
        ''', [nava])
        nava_rows = _rows_to_dicts(cursor, ['nr_lista','id_lista','locatie','lungime','nr_cabluri',
                                        'data','tragator','trimis','data_trimisa','nava','dosar',
                                        'rework','data_rework','ore_rework'])

        cursor.execute('''
            SELECT "Nr Lista","ID Lista","Locatie","Lungime","Nr Cabluri","Data trimisa","Dosar"
            FROM list963 WHERE "Nava" = %s AND "Trimis" = true ORDER BY "Nr Lista" ASC
        ''', [nava])
        trimise = _rows_to_dicts(cursor, ['nr_lista','id_lista','locatie','lungime','nr_cabluri','data_trimisa','dosar'])

        cursor.execute('''
            SELECT "Nr Lista","ID Lista","Locatie","Lungime","Nr Cabluri","Rework"
            FROM list963 WHERE "Nava" = %s AND "Trimis" = false ORDER BY "Nr Lista" ASC
        ''', [nava])
        hala = _rows_to_dicts(cursor, ['nr_lista','id_lista','locatie','lungime','nr_cabluri','rework'])

        cursor.execute('''
            SELECT "Nr Lista","ID Lista","Locatie","Lungime","Nr Cabluri","Data Rework","Ore Rework"
            FROM list963 WHERE "Nava" = %s AND "Rework" = true ORDER BY "Nr Lista" ASC
        ''', [nava])
        rework = _rows_to_dicts(cursor, ['nr_lista','id_lista','locatie','lungime','nr_cabluri','data_rework','ore_rework'])

        cursor.execute('''
            SELECT "Nava", SUM("Lungime"), SUM("Nr Cabluri"), COUNT(*)
            FROM list963 GROUP BY "Nava" ORDER BY "Nava" ASC
        ''')
        all_time = _rows_to_dicts(cursor, ['nava','lungime_totala','cabluri_total','numar_liste'])

    def _sum(lst, key):
        return sum(float(r[key] or 0) for r in lst)

    return render(request, 'stoc/rapoarte.html', {
        'azi': azi,
        'azi_sumar': {'metri': int(_sum(azi,'lungime')), 'cabluri': int(_sum(azi,'nr_cabluri')), 'liste': len(azi)},
        'sapt_sumar': sapt_sumar,
        'zile': zile,
        'luna_sumar': luna_sumar,
        'saptamani': saptamani,
        'nava': nava_rows,
        'nava_sumar': {'metri': int(_sum(nava_rows,'lungime')), 'cabluri': int(_sum(nava_rows,'nr_cabluri')), 'liste': len(nava_rows)},
        'trimise': trimise,
        'trimise_sumar': {'metri': int(_sum(trimise,'lungime')), 'cabluri': int(_sum(trimise,'nr_cabluri')), 'liste': len(trimise)},
        'hala': hala,
        'hala_sumar': {'metri': int(_sum(hala,'lungime')), 'cabluri': int(_sum(hala,'nr_cabluri')), 'liste': len(hala)},
        'rework': rework,
        'rework_sumar': {'metri': int(_sum(rework,'lungime')), 'cabluri': int(_sum(rework,'nr_cabluri')), 'liste': len(rework)},
        'all_time': all_time,
        'all_time_sumar': {'metri': int(_sum(all_time,'lungime_totala')), 'cabluri': int(_sum(all_time,'cabluri_total')), 'liste': int(_sum(all_time,'numar_liste'))},
    })


# ─────────────────────────────────────────────────────────────────────────────
# Super Vizualizare
# ─────────────────────────────────────────────────────────────────────────────

OPERATORS_SQL = {"=": "=", "≠": "!=", ">": ">", "<": "<", "≥": ">=", "≤": "<="}

def _build_query(conditions, vis_cols, sort_field, sort_dir):
    col_sql = ", ".join([f'"{c}"' for c in vis_cols])
    where_parts, params = [], []

    for cond in conditions:
        field = cond.get('field', 'Nava')
        op = cond.get('op', '=')
        val = cond.get('val', '')
        ft = FIELD_TYPES.get(field, 'text')

        col = f'"{field}"'

        if op == 'conține':
            where_parts.append(f'{col} ILIKE %s')
            params.append(f'%{val}%')
        elif op in ('= True', '= False'):
            where_parts.append(f'{col} IS {"TRUE" if op == "= True" else "FALSE"}')
        elif ft == 'numeric' and val:
            try:
                where_parts.append(f'{col} {OPERATORS_SQL.get(op,"=")} %s')
                params.append(float(val.replace(',', '.')))
            except ValueError:
                pass
        elif val:
            where_parts.append(f'{col} {OPERATORS_SQL.get(op,"=")} %s')
            params.append(val)

    where_sql = f'WHERE {" AND ".join(where_parts)}' if where_parts else ''
    order_sql = f'ORDER BY "{sort_field}" {sort_dir}' if sort_field else ''
    query = f'SELECT "ID", {col_sql} FROM list963 {where_sql} {order_sql}'
    return query, params


@_full_access_required
def superviz(request):
    DEFAULT_COLS = ["Nr Lista", "ID Lista", "Lungime", "Nr Cabluri", "Data"]

    if request.method == 'POST':
        action = request.POST.get('action', 'query')

        if action == 'edit':
            ids = [int(i) for i in request.POST.getlist('selected_ids')]
            updates = {}
            field_map = {
                'nr_lista': 'Nr Lista', 'id_lista': 'ID Lista', 'locatie': 'Locatie',
                'dosar': 'Dosar', 'tragator': 'Tragator',
            }
            for key, col in field_map.items():
                val = request.POST.get(key, '').strip()
                if val:
                    updates[col] = val
            for key, col in [('lungime', 'Lungime'), ('ore_rework', 'Ore Rework')]:
                val = request.POST.get(key, '').strip()
                if val:
                    try:
                        updates[col] = float(val.replace(',', '.'))
                    except ValueError:
                        pass
            val = request.POST.get('nr_cabluri', '').strip()
            if val:
                try:
                    updates['Nr Cabluri'] = int(val)
                except ValueError:
                    pass
            for key, col in [('data', 'Data'), ('data_trimisa', 'Data trimisa'), ('data_rework', 'Data Rework')]:
                val = request.POST.get(key, '').strip()
                if val:
                    updates[col] = val
            for key, col in [('trimis', 'Trimis'), ('rework', 'Rework')]:
                val = request.POST.get(key, '')
                if val in ('True', 'False'):
                    updates[col] = val == 'True'
            if updates and ids:
                try:
                    set_parts = [f'"{col}" = %s' for col in updates]
                    params = list(updates.values()) + [ids]
                    with connection.cursor() as cursor:
                        cursor.execute(
                            f'UPDATE list963 SET {", ".join(set_parts)} WHERE "ID" = ANY(%s::int[])',
                            params
                        )
                    messages.success(request, f"{len(ids)} înregistrări actualizate.")
                except Exception as e:
                    messages.error(request, _db_error_msg(e))
            elif not updates:
                messages.warning(request, "Niciun câmp completat pentru editare.")
            return redirect('stoc:superviz')

        if action == 'delete':
            ids = [int(i) for i in request.POST.getlist('selected_ids')]
            if ids:
                try:
                    with connection.cursor() as cursor:
                        cursor.execute('DELETE FROM list963 WHERE "ID" = ANY(%s::int[])', [ids])
                    messages.success(request, f"{len(ids)} înregistrări șterse.")
                except Exception as e:
                    messages.error(request, _db_error_msg(e))
            return redirect('stoc:superviz')

    # Build conditions from GET params
    conditions = []
    i = 0
    while True:
        field = request.GET.get(f'cf_{i}')
        if field is None:
            break
        conditions.append({
            'field': field,
            'op': request.GET.get(f'co_{i}', '='),
            'val': request.GET.get(f'cv_{i}', ''),
        })
        i += 1

    if not conditions:
        nava = int(request.session.get('last_nava', DEFAULT_SHIP))
        conditions = [{'field': 'Nava', 'op': '=', 'val': str(nava)}]

    vis_cols_param = request.GET.get('cols', '')
    vis_cols = vis_cols_param.split(',') if vis_cols_param else DEFAULT_COLS
    vis_cols = [c for c in vis_cols if c in ALL_FIELDS] or DEFAULT_COLS

    sort_field = request.GET.get('sort_field', 'Nr Lista')
    if sort_field not in ALL_FIELDS:
        sort_field = 'Nr Lista'
    sort_dir = 'DESC' if request.GET.get('sort_dir', 'ASC') == 'DESC' else 'ASC'

    query, params = _build_query(conditions, vis_cols, sort_field, sort_dir)

    results = []
    headers = ['ID'] + vis_cols
    try:
        with connection.cursor() as cursor:
            cursor.execute(query, params)
            results = [dict(zip(headers, [_fmt_val(v) for v in row])) for row in cursor.fetchall()]
    except Exception as e:
        messages.error(request, f"Eroare interogare: {e}")

    total_m = sum(float(r.get('Lungime') or 0) for r in results)
    total_c = sum(int(r.get('Nr Cabluri') or 0) for r in results)

    return render(request, 'stoc/superviz.html', {
        'results': results,
        'headers': vis_cols,
        'conditions': conditions,
        'vis_cols': vis_cols,
        'all_fields': ALL_FIELDS,
        'sort_field': sort_field,
        'sort_dir': sort_dir,
        'total_m': int(total_m),
        'total_c': total_c,
        'total_l': len(results),
    })


# ─────────────────────────────────────────────────────────────────────────────
# Export Excel
# ─────────────────────────────────────────────────────────────────────────────

@_full_access_required
def export_borderou(request):
    nava = int(request.session.get('last_nava', DEFAULT_SHIP))
    data_exp = request.GET.get('data', str(date.today()))

    with connection.cursor() as cursor:
        cursor.execute('''
            SELECT "Nr Lista","Locatie","Lungime","Nr Cabluri","Dosar"
            FROM list963
            WHERE "Data trimisa" = %s AND "Nava" = %s
            ORDER BY "Dosar" ASC, "Locatie" ASC, "Nr Lista" ASC
        ''', [data_exp, nava])
        rows = cursor.fetchall()

    if not rows:
        messages.warning(request, "Nu există expedieri pentru această dată.")
        return redirect('stoc:expeditie')

    total_m = sum(float(r[2] or 0) for r in rows)
    total_c = sum(int(r[3] or 0) for r in rows)
    summary = {'metri': int(total_m), 'cabluri': int(total_c), 'liste': len(rows)}

    title = f"Borderou Expeditie Nava {nava} {data_exp}"
    headers = ['Nr Lista', 'Locatie', 'Lungime', 'Nr Cabluri', 'Dosar']
    buffer = _make_excel(rows, headers, title, summary)

    filename = f"Borderou_{nava}_{data_exp}"
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


@login_required
def export_excel(request, tip):
    nava = int(request.session.get('last_nava', DEFAULT_SHIP))
    tip_map = {
        'azi':    (f'Raport_Azi_{nava}',    f'SELECT "Nr Lista","Locatie","Lungime","Nr Cabluri" FROM list963 WHERE "Data" = %s AND "Nava"=%s ORDER BY "Locatie","Nr Lista"', ['Nr Lista','Locatie','Lungime','Nr Cabluri']),
        'nava':   (f'Registru_{nava}',       f'SELECT "Nr Lista","ID Lista","Locatie","Lungime","Nr Cabluri","Data","Tragator","Trimis","Dosar" FROM list963 WHERE "Nava"=%s ORDER BY "Nr Lista"', ['Nr Lista','ID Lista','Locatie','Lungime','Nr Cabluri','Data','Tragator','Trimis','Dosar']),
        'trimise':(f'Liste_Trimise_{nava}',  f'SELECT "Nr Lista","ID Lista","Locatie","Lungime","Nr Cabluri","Data trimisa","Dosar" FROM list963 WHERE "Nava"=%s AND "Trimis"=true ORDER BY "Nr Lista"', ['Nr Lista','ID Lista','Locatie','Lungime','Nr Cabluri','Data trimisa','Dosar']),
        'hala':   (f'Stoc_Hala_{nava}',     f'SELECT "Nr Lista","ID Lista","Locatie","Lungime","Nr Cabluri" FROM list963 WHERE "Nava"=%s AND "Trimis"=false ORDER BY "Nr Lista"', ['Nr Lista','ID Lista','Locatie','Lungime','Nr Cabluri']),
        'rework': (f'Rework_{nava}',         f'SELECT "Nr Lista","ID Lista","Locatie","Lungime","Nr Cabluri","Data Rework","Ore Rework" FROM list963 WHERE "Nava"=%s AND "Rework"=true ORDER BY "Nr Lista"', ['Nr Lista','ID Lista','Locatie','Lungime','Nr Cabluri','Data Rework','Ore Rework']),
    }

    if tip not in tip_map:
        return HttpResponse("Export invalid", status=400)

    filename, query, cols = tip_map[tip]
    params = [date.today(), nava] if tip == 'azi' else [nava]
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    # Compute summary: lungime at col index 2 (0-based), nr_cabluri at 3 — works for all tip_map queries
    try:
        l_idx = cols.index('Lungime')
        c_idx = cols.index('Nr Cabluri')
        summary = {
            'metri': int(sum(float(r[l_idx] or 0) for r in rows)),
            'cabluri': int(sum(int(r[c_idx] or 0) for r in rows)),
            'liste': len(rows),
        }
    except (ValueError, IndexError):
        summary = None

    title = filename.replace('_', ' ')
    buffer = _make_excel(rows, cols, title, summary)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response
